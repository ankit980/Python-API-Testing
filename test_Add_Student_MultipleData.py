import requests
import json
import jsonpath
import openpyxl

def test_Add_new_data():
    api_url="http://thetestingworldapi.com/api/studentsDetails"
    file=open("C:/Users/Ankit/Desktop/API/AddmultipleData.json",'r')
    json_request=json.loads(file.read())
    wk=openpyxl.load_workbook("C:/Users/Ankit/Desktop/API/testData.xlsx")
    sh=wk['Sheet1']
    rows=sh.max_row
    for i in range(2,rows+1):
        cell_first_name=sh.cell(row=i,column=1)
        cell_mid_name=sh.cell(row=i,column=2)
        cell_last_name=sh.cell(row=i,column=3)
        cell_dob=sh.cell(row=i,column=4)

        json_request['first_name']=cell_first_name.value
        json_request['mid_name']=cell_mid_name.value
        json_request['last_name']=cell_last_name.value
        json_request['dob']=cell_dob.value

        response=requests.post(api_url,json_request)
        print(response.status_code)
        print(response.text)